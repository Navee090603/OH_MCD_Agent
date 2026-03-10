"""Excel writing helpers for OH MCD submission reporting."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

LOGGER = logging.getLogger(__name__)


class ExcelWriter:
    """Append query results into a new worksheet for each run."""

    def __init__(self, workbook_path: str, worksheet_prefix: str) -> None:
        self.workbook_path = Path(workbook_path)
        self.worksheet_prefix = worksheet_prefix

    def _ensure_workbook(self) -> None:
        if self.workbook_path.exists():
            return

        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        default_sheet = workbook.active
        default_sheet.title = "README"
        default_sheet["A1"] = "Auto-generated workbook for OH MCD submission counts"
        workbook.save(self.workbook_path)
        LOGGER.info("Created new workbook at %s", self.workbook_path)

    def write_results(self, dataframe: pd.DataFrame, run_dt: datetime) -> str:
        """Create a new worksheet and write DataFrame contents.

        Returns created worksheet name.
        """
        self._ensure_workbook()
        worksheet_name = f"{self.worksheet_prefix}_{run_dt:%Y_%m_%d}"

        workbook = load_workbook(self.workbook_path)
        if worksheet_name in workbook.sheetnames:
            base_name = worksheet_name
            suffix = 1
            while worksheet_name in workbook.sheetnames:
                worksheet_name = f"{base_name}_{suffix}"
                suffix += 1

        sheet = workbook.create_sheet(title=worksheet_name)

        for row_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True)

        sheet.freeze_panes = "A2"
        workbook.save(self.workbook_path)
        workbook.close()

        LOGGER.info("Wrote %s rows to worksheet %s", len(dataframe), worksheet_name)
        return worksheet_name
